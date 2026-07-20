from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config import RELATORIOS_DIR


def exportar_excel(imoveis):

    RELATORIOS_DIR.mkdir(exist_ok=True)

    arquivo = RELATORIOS_DIR / f"captacao_{date.today():%Y-%m-%d}.xlsx"

    wb = Workbook()


    ws = wb.active


    ws.title = "Oportunidades"



    cabecalho = [

        "Portal",

        "Tipo",

        "Título",

        "Bairro",

        "Cidade",

        "Valor",

        "Quartos",

        "Área Privativa",

        "Área Total",

        "Pontuação",

        "Classificação",

        "Link"

    ]



    ws.append(

        cabecalho

    )



    for celula in ws[1]:

        celula.font = Font(

            bold=True

        )

        celula.fill = PatternFill(

            "solid",

            fgColor="FFD966"

        )



    for imovel in imoveis:


        ws.append([

            imovel.portal,

            imovel.tipo,

            imovel.titulo,

            imovel.bairro,

            imovel.cidade,

            imovel.valor,

            imovel.quartos,

            getattr(

                imovel,

                "area_privativa",

                ""

            ),

            getattr(

                imovel,

                "area_total",

                ""

            ),

            getattr(

                imovel,

                "pontuacao",

                0

            ),

            getattr(

                imovel,

                "classificacao",

                ""

            ),

            imovel.link

        ])



    for coluna in ws.columns:

        tamanho = max(

            len(str(c.value))

            if c.value else 0

            for c in coluna

        )


        ws.column_dimensions[

            coluna[0].column_letter

        ].width = tamanho + 5



    wb.save(

        arquivo

    )


    print(

        "Excel criado:",

        arquivo

    )

    return arquivo